#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
更新 Excel 文件：将 TSV 数据追加到现有 Excel 工作簿
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os

# 文件路径
excel_file = "美泰三温四区整机软件方案问题点汇总.xlsx"
tsv_files = {
    "完成情况": "excel/三温四区_完成情况.tsv",
    "难点风险": "excel/三温四区_难点风险.tsv",
    "下一步": "excel/三温四区_下一步_1-2周.tsv"
}

def read_tsv(filepath):
    """读取 TSV 文件并返回二维列表"""
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    data = [line.strip().split('\t') for line in lines]
    return data

def append_to_sheet(ws, data, start_row=None):
    """将数据追加到工作表"""
    if start_row is None:
        start_row = ws.max_row + 2  # 空一行
    
    # 写入数据
    for i, row_data in enumerate(data):
        for j, cell_value in enumerate(row_data):
            cell = ws.cell(row=start_row + i, column=j + 1, value=cell_value)
            
            # 设置表头样式
            if i == 0:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    return start_row + len(data)

def main():
    # 检查 Excel 文件是否存在
    if not os.path.exists(excel_file):
        print(f"错误：找不到文件 {excel_file}")
        return
    
    # 打开工作簿
    wb = openpyxl.load_workbook(excel_file)
    
    # 获取或创建工作表
    sheet_name = "三温四区项目汇总"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    
    # 清空现有内容（可选）
    # ws.delete_rows(1, ws.max_row)
    
    current_row = 1
    
    # 添加标题
    title_cell = ws.cell(row=current_row, column=1, value="三温四区项目完成情况与难点汇总")
    title_cell.font = Font(bold=True, size=14)
    current_row += 2
    
    # 1. 完成情况
    section_cell = ws.cell(row=current_row, column=1, value="一、完成情况（按模块/里程碑）")
    section_cell.font = Font(bold=True, size=12, color="1F4E78")
    current_row += 1
    
    data = read_tsv(tsv_files["完成情况"])
    current_row = append_to_sheet(ws, data, current_row)
    current_row += 2
    
    # 2. 难点风险
    section_cell = ws.cell(row=current_row, column=1, value="二、难点 / 风险点（含影响与对策）")
    section_cell.font = Font(bold=True, size=12, color="1F4E78")
    current_row += 1
    
    data = read_tsv(tsv_files["难点风险"])
    current_row = append_to_sheet(ws, data, current_row)
    current_row += 2
    
    # 3. 下一步
    section_cell = ws.cell(row=current_row, column=1, value="三、下一步（1~2 周）")
    section_cell.font = Font(bold=True, size=12, color="1F4E78")
    current_row += 1
    
    data = read_tsv(tsv_files["下一步"])
    current_row = append_to_sheet(ws, data, current_row)
    
    # 调整列宽
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 30
    
    # 保存
    wb.save(excel_file)
    print(f"✓ 已更新 {excel_file}")
    print(f"  - 工作表：{sheet_name}")
    print(f"  - 总行数：{current_row}")

if __name__ == "__main__":
    main()
